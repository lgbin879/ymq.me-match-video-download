#!/usr/local/bin/python3
#usage: python3 f64_fc_fw_srds_extra_from_excel.py EDCS-Hindon_SD_HSD_Trace_Details_and_SerDes_Settings_20200520.xlsx

import sys
import copy
import openpyxl
from pymysql import *

param_col_dict = {
'name' :    1,	
'gender' :  2,	
'age' :     3,	
'birthday' :4,	
'phone' :   5,
'club':	    6,
'id_num' :  7
}

sql_sequence = """
create table personal_info(
    id int unsigned primary key auto_increment not null,
    name varchar(10) not null,
    gender varchar(4) ,
    age int unsigned ,
    birthday varchar(20) ,
    phone varchar(20) ,
    club varchar(20) ,
    id_num varchar(20)  
);

--- create club table from personal_info group by club
create table if not exists club_name( 
     id int unsigned primary key auto_increment not null, 
     name varchar(50) default NULL 
 )select club as name from personal_info group by club;

 select * from personal_info as p inner join club_name as c on p.club = c.name;

 update goods as g inner join goods_cates as c on g.cate_name = c.name set g.cate_name = c.id;

 alter table goods change cate_name cate_id int unsigned not null; 

 select p.*, c.name from personal_info as p left join club_name as c on p.club_id=c.id;
 --- 视图
 create view v_p_info 
 as select p.*, c.name as club_name from personal_info as p left join club_name as c on p.club_id=c.id; 

 alter table club_name change count number int unsigned;

 --- get number from personal_info table
 update club_name as c set `number`=(select count(*) from personal_info where club_id=c.id);
 update club_name as c set `male_num`=(select count(*) from personal_info as p where p.club_id=c.id and p.gender='男');
 update club_name as c set `female_num`=(select count(*) from personal_info as p where p.club_id=c.id and p.gender='女');

# 开启事务
begin;
# 或者
start transaction;

commit; or rollback;

# Backup
mysqldump -uroot -pmysql ymq_me > ymq_me.sql

# Import
mysql -uroot -p ymq_me < ymq_me.sql

"""



class YMQME(object):
    def __init__(self):
        # 创建connection连接  连接对象
        self.conn = connect(host="localhost", 
                            port=3306, 
                            user='root', 
                            password='mysql', 
                            database='ymq_me', 
                            charset='utf8')
        # 获取Cursor对象  游标对象
        self.cursor = self.conn.cursor()

    def __del__(self):
        # 关闭cursor对象
        self.cursor.close()
        self.conn.close()

    def execute_sql(self, sql):
        self.cursor.execute(sql)

    def insert_data(self, data_tuple):
        sql = """insert into personal_info values {0}""".format(data_tuple)
        self.execute_sql(sql)

    def insert_data_from_xsl(self, info_sheet):
        for row in range(2, info_sheet.max_row+1):
            name = str(info_sheet.cell(row, param_col_dict['name']).value)
            gender = str(info_sheet.cell(row, param_col_dict['gender']).value)
            age = int(info_sheet.cell(row, param_col_dict['age']).value)
            birthday = str(info_sheet.cell(row, param_col_dict['birthday']).value)
            phone = str(info_sheet.cell(row, param_col_dict['phone']).value)
            club = str(info_sheet.cell(row, param_col_dict['club']).value)
            id_num = str(info_sheet.cell(row, param_col_dict['id_num']).value)

            data_list = [0, name, gender, age, birthday, phone, club, id_num]
            #data_list = [0, name, gender, age, phone, club, birthday, id_num]
            data_tuple = tuple(data_list)
            sql = """insert into personal_info values {0}""".format(data_tuple)
            #sql = """insert into personinfo_personinfo values {0}""".format(data_tuple)
            print("----->{0}<------".format(sql))
            self.insert_data(data_tuple)

        self.conn.commit() 


def main():
    argc = len(sys.argv)
    if argc < 2:
        print("## Error : no xlsx file input")
        print("## Usage : python3 %s file_name" % sys.argv[0])
        return True

    file_name = sys.argv[1]
    wb = openpyxl.load_workbook(file_name)
    info_sheet = wb['sheet1']

    ymq_me= YMQME()
    ymq_me.insert_data_from_xsl(info_sheet)

if __name__ == "__main__":
    main()
